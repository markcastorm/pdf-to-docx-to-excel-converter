#!/usr/bin/env python3
"""
JGBF Excel Parser - Fixed Version
Reads extracted Excel files and generates JGBF_DATA output in the exact format required.
Processes data from tittle3.py output and creates standardized time series data.
"""

import os
import sys
from pathlib import Path
import logging
import re
from typing import Dict, List, Tuple, Optional
from datetime import datetime
import pandas as pd

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
except ImportError as e:
    print("Missing required packages. Install with:")
    print("pip install openpyxl pandas")
    sys.exit(1)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class JGBFParser:
    """Parses Excel files from tittle3.py and generates JGBF_DATA output."""
    
    def __init__(self, input_folder: str = "extracted_data", output_folder: str = "parsed_output"):
        self.input_folder = Path(input_folder)
        self.output_folder = Path(output_folder)
        self.output_folder.mkdir(exist_ok=True)
        
        # Instrument mapping from subtitle to code (flexible matching)
        self.instrument_mapping = {
            "JGB(10-year) Futures": "JGB10YEARFUTURES",
            "mini-10-year JGB Futures": "MINI10YEARJGBFUTURESCASHSETTLED",  # Matches both (Cash-Settled) and（Cash-Settled)
            "mini-20-year JGB Futures": "MINI20YEARJGBFUTURES",
            "3-Month TONA Futures": "3MONTHTONAFUTURES"
        }
        
        # Category mapping for Main Summary tables
        self.main_summary_categories = {
            "自己取引計": "PROPRIETARY",
            "委託取引計": "BROKERAGE", 
            "自己委託合計": "TOTAL"
        }
        
        # Category mapping for Brokerage Breakdown tables  
        self.brokerage_categories = {
            "法人計": "INSTITUTIONS",
            "個人計": "INDIVIDUALS",
            "海外投資家計": "FOREIGNERS",
            "証券会社": "SECURITIES_COS"
        }
        
        # Subcategory mapping
        self.subcategories = {
            "売り": "SALES",
            "買い": "PURCHASES"
        }
        
        # Week mapping for date conversion
        self.week_mapping = {
            "MARCH WEEK 1": "2025-09",
            "MARCH WEEK 2": "2025-10", 
            "MARCH WEEK 3": "2025-11",
            "MARCH WEEK 4": "2025-12"
        }
        
        logger.info(f"🚀 JGBF Parser initialized")
        logger.info(f"  • Input folder: {self.input_folder}")
        logger.info(f"  • Output folder: {self.output_folder}")

    def get_template_columns(self):
        """Return the exact column structure from the JGBF_DATA template."""
        # This is the exact column order and structure from your template file
        columns = [
            # JGB 10-year Futures - Total, Proprietary & Brokerage
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.JGB10YEARFUTURES.TRADINGVALUE.PROPRIETARY.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Total, Proprietary ＆ Brokerage, Trading Value, Proprietary, Sales, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.JGB10YEARFUTURES.TRADINGVALUE.PROPRIETARY.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Total, Proprietary ＆ Brokerage, Trading Value, Proprietary, Sales, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.JGB10YEARFUTURES.TRADINGVALUE.PROPRIETARY.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Total, Proprietary ＆ Brokerage, Trading Value, Proprietary, Purchases, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.JGB10YEARFUTURES.TRADINGVALUE.PROPRIETARY.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Total, Proprietary ＆ Brokerage, Trading Value, Proprietary, Purchases, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.JGB10YEARFUTURES.TRADINGVALUE.BROKERAGE.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Total, Proprietary ＆ Brokerage, Trading Value, Brokerage, Sales, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.JGB10YEARFUTURES.TRADINGVALUE.BROKERAGE.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Total, Proprietary ＆ Brokerage, Trading Value, Brokerage, Sales, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.JGB10YEARFUTURES.TRADINGVALUE.BROKERAGE.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Total, Proprietary ＆ Brokerage, Trading Value, Brokerage, Purchases, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.JGB10YEARFUTURES.TRADINGVALUE.BROKERAGE.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Total, Proprietary ＆ Brokerage, Trading Value, Brokerage, Purchases, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.JGB10YEARFUTURES.TRADINGVALUE.TOTAL.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Total, Proprietary ＆ Brokerage, Trading Value, Total, Sales, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.JGB10YEARFUTURES.TRADINGVALUE.TOTAL.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Total, Proprietary ＆ Brokerage, Trading Value, Total, Sales, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.JGB10YEARFUTURES.TRADINGVALUE.TOTAL.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Total, Proprietary ＆ Brokerage, Trading Value, Total, Purchases, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.JGB10YEARFUTURES.TRADINGVALUE.TOTAL.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Total, Proprietary ＆ Brokerage, Trading Value, Total, Purchases, Balance'
            },
            
            # JGB 10-year Futures - Brokerage Breakdown
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.JGB10YEARFUTURES.TRADINGVALUE.INSTITUTIONS.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Breakdown of Brokerage, Trading Value, Institutions, Sales, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.JGB10YEARFUTURES.TRADINGVALUE.INSTITUTIONS.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Breakdown of Brokerage, Trading Value, Institutions, Sales, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.JGB10YEARFUTURES.TRADINGVALUE.INSTITUTIONS.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Breakdown of Brokerage, Trading Value, Institutions, Purchases, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.JGB10YEARFUTURES.TRADINGVALUE.INSTITUTIONS.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Breakdown of Brokerage, Trading Value, Institutions, Purchases, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.JGB10YEARFUTURES.TRADINGVALUE.INDIVIDUALS.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Breakdown of Brokerage, Trading Value, Individuals, Sales, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.JGB10YEARFUTURES.TRADINGVALUE.INDIVIDUALS.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Breakdown of Brokerage, Trading Value, Individuals, Sales, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.JGB10YEARFUTURES.TRADINGVALUE.INDIVIDUALS.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Breakdown of Brokerage, Trading Value, Individuals, Purchases, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.JGB10YEARFUTURES.TRADINGVALUE.INDIVIDUALS.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Breakdown of Brokerage, Trading Value, Individuals, Purchases, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.JGB10YEARFUTURES.TRADINGVALUE.FOREIGNERS.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Breakdown of Brokerage, Trading Value, Foreigners, Sales, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.JGB10YEARFUTURES.TRADINGVALUE.FOREIGNERS.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Breakdown of Brokerage, Trading Value, Foreigners, Sales, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.JGB10YEARFUTURES.TRADINGVALUE.FOREIGNERS.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Breakdown of Brokerage, Trading Value, Foreigners, Purchases, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.JGB10YEARFUTURES.TRADINGVALUE.FOREIGNERS.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Breakdown of Brokerage, Trading Value, Foreigners, Purchases, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.JGB10YEARFUTURES.TRADINGVALUE.SECURITIES_COS.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Breakdown of Brokerage, Trading Value, Securities Cos., Sales, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.JGB10YEARFUTURES.TRADINGVALUE.SECURITIES_COS.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Breakdown of Brokerage, Trading Value, Securities Cos., Sales, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.JGB10YEARFUTURES.TRADINGVALUE.SECURITIES_COS.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Breakdown of Brokerage, Trading Value, Securities Cos., Purchases, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.JGB10YEARFUTURES.TRADINGVALUE.SECURITIES_COS.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, JGB(10-year) Futures, Breakdown of Brokerage, Trading Value, Securities Cos., Purchases, Balance'
            },
            
            # Mini-10-year JGB Futures (Cash-Settled) - Total, Proprietary & Brokerage
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.PROPRIETARY.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Total, Proprietary ＆ Brokerage, Trading Value, Proprietary, Sales, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.PROPRIETARY.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Total, Proprietary ＆ Brokerage, Trading Value, Proprietary, Sales, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.PROPRIETARY.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Total, Proprietary ＆ Brokerage, Trading Value, Proprietary, Purchases, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.PROPRIETARY.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Total, Proprietary ＆ Brokerage, Trading Value, Proprietary, Purchases, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.BROKERAGE.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Total, Proprietary ＆ Brokerage, Trading Value, Brokerage, Sales, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.BROKERAGE.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Total, Proprietary ＆ Brokerage, Trading Value, Brokerage, Sales, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.BROKERAGE.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Total, Proprietary ＆ Brokerage, Trading Value, Brokerage, Purchases, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.BROKERAGE.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Total, Proprietary ＆ Brokerage, Trading Value, Brokerage, Purchases, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.TOTAL.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Total, Proprietary ＆ Brokerage, Trading Value, Total, Sales, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.TOTAL.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Total, Proprietary ＆ Brokerage, Trading Value, Total, Sales, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.TOTAL.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Total, Proprietary ＆ Brokerage, Trading Value, Total, Purchases, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.TOTAL.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Total, Proprietary ＆ Brokerage, Trading Value, Total, Purchases, Balance'
            },
            
            # Mini-10-year JGB Futures (Cash-Settled) - Brokerage Breakdown
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.INSTITUTIONS.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Breakdown of Brokerage, Trading Value, Institutions, Sales, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.INSTITUTIONS.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Breakdown of Brokerage, Trading Value, Institutions, Sales, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.INSTITUTIONS.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Breakdown of Brokerage, Trading Value, Institutions, Purchases, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.INSTITUTIONS.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Breakdown of Brokerage, Trading Value, Institutions, Purchases, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.INDIVIDUALS.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Breakdown of Brokerage, Trading Value, Individuals, Sales, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.INDIVIDUALS.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Breakdown of Brokerage, Trading Value, Individuals, Sales, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.INDIVIDUALS.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Breakdown of Brokerage, Trading Value, Individuals, Purchases, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.INDIVIDUALS.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Breakdown of Brokerage, Trading Value, Individuals, Purchases, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.FOREIGNERS.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Breakdown of Brokerage, Trading Value, Foreigners, Sales, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.FOREIGNERS.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Breakdown of Brokerage, Trading Value, Foreigners, Sales, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.FOREIGNERS.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Breakdown of Brokerage, Trading Value, Foreigners, Purchases, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.FOREIGNERS.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Breakdown of Brokerage, Trading Value, Foreigners, Purchases, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.SECURITIES_COS.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Breakdown of Brokerage, Trading Value, Securities Cos., Sales, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.SECURITIES_COS.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Breakdown of Brokerage, Trading Value, Securities Cos., Sales, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.SECURITIES_COS.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Breakdown of Brokerage, Trading Value, Securities Cos., Purchases, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI10YEARJGBFUTURESCASHSETTLED.TRADINGVALUE.SECURITIES_COS.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-10-year JGB Futures (Cash-Settled), Breakdown of Brokerage, Trading Value, Securities Cos., Purchases, Balance'
            },
            
            # Mini-20-year JGB Futures - Total, Proprietary & Brokerage
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI20YEARJGBFUTURES.TRADINGVALUE.PROPRIETARY.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Total, Proprietary ＆ Brokerage, Trading Value, Proprietary, Sales, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI20YEARJGBFUTURES.TRADINGVALUE.PROPRIETARY.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Total, Proprietary ＆ Brokerage, Trading Value, Proprietary, Sales, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI20YEARJGBFUTURES.TRADINGVALUE.PROPRIETARY.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Total, Proprietary ＆ Brokerage, Trading Value, Proprietary, Purchases, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI20YEARJGBFUTURES.TRADINGVALUE.PROPRIETARY.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Total, Proprietary ＆ Brokerage, Trading Value, Proprietary, Purchases, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI20YEARJGBFUTURES.TRADINGVALUE.BROKERAGE.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Total, Proprietary ＆ Brokerage, Trading Value, Brokerage, Sales, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI20YEARJGBFUTURES.TRADINGVALUE.BROKERAGE.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Total, Proprietary ＆ Brokerage, Trading Value, Brokerage, Sales, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI20YEARJGBFUTURES.TRADINGVALUE.BROKERAGE.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Total, Proprietary ＆ Brokerage, Trading Value, Brokerage, Purchases, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI20YEARJGBFUTURES.TRADINGVALUE.BROKERAGE.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Total, Proprietary ＆ Brokerage, Trading Value, Brokerage, Purchases, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI20YEARJGBFUTURES.TRADINGVALUE.TOTAL.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Total, Proprietary ＆ Brokerage, Trading Value, Total, Sales, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI20YEARJGBFUTURES.TRADINGVALUE.TOTAL.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Total, Proprietary ＆ Brokerage, Trading Value, Total, Sales, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI20YEARJGBFUTURES.TRADINGVALUE.TOTAL.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Total, Proprietary ＆ Brokerage, Trading Value, Total, Purchases, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.MINI20YEARJGBFUTURES.TRADINGVALUE.TOTAL.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Total, Proprietary ＆ Brokerage, Trading Value, Total, Purchases, Balance'
            },
            
            # Mini-20-year JGB Futures - Brokerage Breakdown
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI20YEARJGBFUTURES.TRADINGVALUE.INSTITUTIONS.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Breakdown of Brokerage, Trading Value, Institutions, Sales, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI20YEARJGBFUTURES.TRADINGVALUE.INSTITUTIONS.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Breakdown of Brokerage, Trading Value, Institutions, Sales, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI20YEARJGBFUTURES.TRADINGVALUE.INSTITUTIONS.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Breakdown of Brokerage, Trading Value, Institutions, Purchases, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI20YEARJGBFUTURES.TRADINGVALUE.INSTITUTIONS.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Breakdown of Brokerage, Trading Value, Institutions, Purchases, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI20YEARJGBFUTURES.TRADINGVALUE.INDIVIDUALS.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Breakdown of Brokerage, Trading Value, Individuals, Sales, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI20YEARJGBFUTURES.TRADINGVALUE.INDIVIDUALS.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Breakdown of Brokerage, Trading Value, Individuals, Sales, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI20YEARJGBFUTURES.TRADINGVALUE.INDIVIDUALS.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Breakdown of Brokerage, Trading Value, Individuals, Purchases, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI20YEARJGBFUTURES.TRADINGVALUE.INDIVIDUALS.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Breakdown of Brokerage, Trading Value, Individuals, Purchases, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI20YEARJGBFUTURES.TRADINGVALUE.FOREIGNERS.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Breakdown of Brokerage, Trading Value, Foreigners, Sales, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI20YEARJGBFUTURES.TRADINGVALUE.FOREIGNERS.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Breakdown of Brokerage, Trading Value, Foreigners, Sales, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI20YEARJGBFUTURES.TRADINGVALUE.FOREIGNERS.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Breakdown of Brokerage, Trading Value, Foreigners, Purchases, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI20YEARJGBFUTURES.TRADINGVALUE.FOREIGNERS.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Breakdown of Brokerage, Trading Value, Foreigners, Purchases, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI20YEARJGBFUTURES.TRADINGVALUE.SECURITIES_COS.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Breakdown of Brokerage, Trading Value, Securities Cos., Sales, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI20YEARJGBFUTURES.TRADINGVALUE.SECURITIES_COS.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Breakdown of Brokerage, Trading Value, Securities Cos., Sales, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI20YEARJGBFUTURES.TRADINGVALUE.SECURITIES_COS.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Breakdown of Brokerage, Trading Value, Securities Cos., Purchases, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.MINI20YEARJGBFUTURES.TRADINGVALUE.SECURITIES_COS.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, mini-20-year JGB Futures, Breakdown of Brokerage, Trading Value, Securities Cos., Purchases, Balance'
            },
            
            # 3-Month TONA Futures - Total, Proprietary & Brokerage
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.3MONTHTONAFUTURES.TRADINGVALUE.PROPRIETARY.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Total, Proprietary ＆ Brokerage, Trading Value, Proprietary, Sales, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.3MONTHTONAFUTURES.TRADINGVALUE.PROPRIETARY.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Total, Proprietary ＆ Brokerage, Trading Value, Proprietary, Sales, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.3MONTHTONAFUTURES.TRADINGVALUE.PROPRIETARY.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Total, Proprietary ＆ Brokerage, Trading Value, Proprietary, Purchases, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.3MONTHTONAFUTURES.TRADINGVALUE.PROPRIETARY.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Total, Proprietary ＆ Brokerage, Trading Value, Proprietary, Purchases, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.3MONTHTONAFUTURES.TRADINGVALUE.BROKERAGE.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Total, Proprietary ＆ Brokerage, Trading Value, Brokerage, Sales, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.3MONTHTONAFUTURES.TRADINGVALUE.BROKERAGE.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Total, Proprietary ＆ Brokerage, Trading Value, Brokerage, Sales, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.3MONTHTONAFUTURES.TRADINGVALUE.BROKERAGE.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Total, Proprietary ＆ Brokerage, Trading Value, Brokerage, Purchases, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.3MONTHTONAFUTURES.TRADINGVALUE.BROKERAGE.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Total, Proprietary ＆ Brokerage, Trading Value, Brokerage, Purchases, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.3MONTHTONAFUTURES.TRADINGVALUE.TOTAL.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Total, Proprietary ＆ Brokerage, Trading Value, Total, Sales, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.3MONTHTONAFUTURES.TRADINGVALUE.TOTAL.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Total, Proprietary ＆ Brokerage, Trading Value, Total, Sales, Balance'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.3MONTHTONAFUTURES.TRADINGVALUE.TOTAL.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Total, Proprietary ＆ Brokerage, Trading Value, Total, Purchases, Value'
            },
            {
                'code': 'JGBF.TOTAL_PROPRIETARY_BROKERAGE.3MONTHTONAFUTURES.TRADINGVALUE.TOTAL.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Total, Proprietary ＆ Brokerage, Trading Value, Total, Purchases, Balance'
            },
            
            # 3-Month TONA Futures - Brokerage Breakdown
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.3MONTHTONAFUTURES.TRADINGVALUE.INSTITUTIONS.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Breakdown of Brokerage, Trading Value, Institutions, Sales, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.3MONTHTONAFUTURES.TRADINGVALUE.INSTITUTIONS.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Breakdown of Brokerage, Trading Value, Institutions, Sales, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.3MONTHTONAFUTURES.TRADINGVALUE.INSTITUTIONS.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Breakdown of Brokerage, Trading Value, Institutions, Purchases, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.3MONTHTONAFUTURES.TRADINGVALUE.INSTITUTIONS.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Breakdown of Brokerage, Trading Value, Institutions, Purchases, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.3MONTHTONAFUTURES.TRADINGVALUE.INDIVIDUALS.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Breakdown of Brokerage, Trading Value, Individuals, Sales, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.3MONTHTONAFUTURES.TRADINGVALUE.INDIVIDUALS.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Breakdown of Brokerage, Trading Value, Individuals, Sales, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.3MONTHTONAFUTURES.TRADINGVALUE.INDIVIDUALS.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Breakdown of Brokerage, Trading Value, Individuals, Purchases, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.3MONTHTONAFUTURES.TRADINGVALUE.INDIVIDUALS.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Breakdown of Brokerage, Trading Value, Individuals, Purchases, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.3MONTHTONAFUTURES.TRADINGVALUE.FOREIGNERS.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Breakdown of Brokerage, Trading Value, Foreigners, Sales, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.3MONTHTONAFUTURES.TRADINGVALUE.FOREIGNERS.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Breakdown of Brokerage, Trading Value, Foreigners, Sales, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.3MONTHTONAFUTURES.TRADINGVALUE.FOREIGNERS.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Breakdown of Brokerage, Trading Value, Foreigners, Purchases, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.3MONTHTONAFUTURES.TRADINGVALUE.FOREIGNERS.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Breakdown of Brokerage, Trading Value, Foreigners, Purchases, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.3MONTHTONAFUTURES.TRADINGVALUE.SECURITIES_COS.SALES.VALUE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Breakdown of Brokerage, Trading Value, Securities Cos., Sales, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.3MONTHTONAFUTURES.TRADINGVALUE.SECURITIES_COS.SALES.BALANCE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Breakdown of Brokerage, Trading Value, Securities Cos., Sales, Balance'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.3MONTHTONAFUTURES.TRADINGVALUE.SECURITIES_COS.PURCHASES.VALUE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Breakdown of Brokerage, Trading Value, Securities Cos., Purchases, Value'
            },
            {
                'code': 'JGBF.BROKERAGE_BREAKDOWN.3MONTHTONAFUTURES.TRADINGVALUE.SECURITIES_COS.PURCHASES.BALANCE.W',
                'description': 'Trading by Type of Investors, 3-Month TONA Futures, Breakdown of Brokerage, Trading Value, Securities Cos., Purchases, Balance'
            }
        ]
        
        return columns

    def extract_instrument_from_subtitle(self, subtitle: str) -> Optional[str]:
        """Extract instrument code from subtitle using enhanced flexible matching."""
        # Clean up subtitle for better matching - handle all parentheses variations
        subtitle_clean = subtitle.replace("（", "(").replace("）", ")").replace("、", ",")
        
        # Enhanced matching with Japanese text recognition
        logger.info(f"🔍 Analyzing subtitle: {subtitle}")
        logger.info(f"🔍 Cleaned subtitle: {subtitle_clean}")
        
        # More flexible matching - check for key patterns
        if "JGB(10-year)" in subtitle_clean or "長期国債先物" in subtitle_clean:
            if "mini" in subtitle_clean.lower() or "ミニ" in subtitle_clean or "現金決済型ミニ" in subtitle_clean:
                logger.info("✅ Matched: MINI10YEARJGBFUTURESCASHSETTLED")
                return "MINI10YEARJGBFUTURESCASHSETTLED"
            else:
                logger.info("✅ Matched: JGB10YEARFUTURES")
                return "JGB10YEARFUTURES"
        elif "mini-20-year" in subtitle_clean.lower() or "20年" in subtitle_clean:
            logger.info("✅ Matched: MINI20YEARJGBFUTURES")
            return "MINI20YEARJGBFUTURES"
        elif "3-Month TONA" in subtitle_clean or "TONA" in subtitle_clean:
            logger.info("✅ Matched: 3MONTHTONAFUTURES")
            return "3MONTHTONAFUTURES"
            
        # Fallback to original mapping for any missed cases
        for key, code in self.instrument_mapping.items():
            if key in subtitle_clean:
                logger.info(f"✅ Fallback matched: {code}")
                return code
                
        logger.warning(f"❌ Could not map subtitle to instrument: {subtitle}")
        return None

    def extract_date_from_filename(self, filename: str) -> Optional[str]:
        """Extract date from filename."""
        # Handle various filename patterns
        filename_upper = filename.upper()
        
        for week_pattern, date_code in self.week_mapping.items():
            if week_pattern in filename_upper:
                return date_code
                
        # Try to extract date patterns like YYYYMMDD
        date_match = re.search(r'(\d{8})', filename)
        if date_match:
            date_str = date_match.group(1)
            try:
                date_obj = datetime.strptime(date_str, '%Y%m%d')
                # Convert to week format (approximate)
                week_num = date_obj.isocalendar()[1]
                return f"{date_obj.year}-{week_num:02d}"
            except ValueError:
                pass
                
        logger.warning(f"Could not extract date from filename: {filename}")
        return "2025-01"  # Default fallback

    def handle_negative_values(self, value: str) -> str:
        """Convert ▲ symbols to negative values."""
        if not value or value == "-" or value == "":
            return ""
        
        value_str = str(value).strip()
        if value_str.startswith("▲"):
            # Remove ▲ and add negative sign
            return "-" + value_str[1:]
        return value_str

    def read_excel_sheet(self, file_path: Path, sheet_name: str) -> Optional[Dict]:
        """Read specific sheet from Excel file."""
        try:
            wb = load_workbook(file_path, data_only=True)
            if sheet_name not in wb.sheetnames:
                return None
                
            ws = wb[sheet_name]
            
            # Extract metadata from first few rows
            title = ws['A1'].value or ""
            subtitle = ws['A2'].value or ""
            table_title = ws['A3'].value or ""
            
            # Extract subtitle content (remove "Subtitle: " prefix)
            if subtitle.startswith("Subtitle: "):
                subtitle = subtitle[10:]
            
            # Extract table title content (remove "Table Title: " prefix)  
            if table_title.startswith("Table Title: "):
                table_title = table_title[13:]
                
            # Read data starting from row 6 (after headers)
            data_rows = []
            for row in ws.iter_rows(min_row=6, values_only=True):
                if row[0]:  # If first column has data
                    data_rows.append(row)
                    
            return {
                'title': title,
                'subtitle': subtitle, 
                'table_title': table_title,
                'data_rows': data_rows
            }
            
        except Exception as e:
            logger.error(f"Error reading sheet {sheet_name} from {file_path}: {e}")
            return None

    def parse_main_summary_table(self, sheet_data: Dict, instrument_code: str, date_code: str) -> List[Dict]:
        """Parse Main Summary table data."""
        results = []
        
        for row in sheet_data['data_rows']:
            if len(row) < 8:
                continue
                
            category_full = str(row[0] or "").strip()
            subcategory = str(row[1] or "").strip()
            value = row[5]  # Column F (金額 Value)
            balance = row[7]  # Column H (差引き Balance)
            
            # Skip if not a valid data row
            if not category_full or not subcategory:
                continue
                
            # Skip total rows (合計 Total)
            if "合計" in subcategory:
                continue
                
            # Extract category
            category = None
            for jp_cat, en_cat in self.main_summary_categories.items():
                if jp_cat in category_full:
                    category = en_cat
                    break
                    
            if not category:
                continue
                
            # Extract subcategory
            subcat = None
            for jp_sub, en_sub in self.subcategories.items():
                if jp_sub in subcategory:
                    subcat = en_sub
                    break
                    
            if not subcat:
                continue
                
            # Process VALUE
            if value is not None:
                value_str = self.handle_negative_values(str(value))
                if value_str:
                    value_code = f"JGBF.TOTAL_PROPRIETARY_BROKERAGE.{instrument_code}.TRADINGVALUE.{category}.{subcat}.VALUE.W"
                    value_desc = f"Trading by Type of Investors, {self.get_instrument_name(instrument_code)}, Total, Proprietary ＆ Brokerage, Trading Value, {category.title()}, {subcat.title()}, Value"
                    
                    results.append({
                        'code': value_code,
                        'description': value_desc,
                        'date': date_code,
                        'value': value_str
                    })
            
            # Process BALANCE  
            if balance is not None:
                balance_str = self.handle_negative_values(str(balance))
                if balance_str:
                    balance_code = f"JGBF.TOTAL_PROPRIETARY_BROKERAGE.{instrument_code}.TRADINGVALUE.{category}.{subcat}.BALANCE.W"
                    balance_desc = f"Trading by Type of Investors, {self.get_instrument_name(instrument_code)}, Total, Proprietary ＆ Brokerage, Trading Value, {category.title()}, {subcat.title()}, Balance"
                    
                    results.append({
                        'code': balance_code,
                        'description': balance_desc, 
                        'date': date_code,
                        'value': balance_str
                    })
                    
        return results

    def parse_brokerage_breakdown_table(self, sheet_data: Dict, instrument_code: str, date_code: str) -> List[Dict]:
        """Parse Brokerage Breakdown table data."""
        results = []
        
        for row in sheet_data['data_rows']:
            if len(row) < 8:
                continue
                
            category_full = str(row[0] or "").strip()
            subcategory = str(row[1] or "").strip() 
            value = row[5]  # Column F (金額 Value)
            balance = row[7]  # Column H (差引き Balance)
            
            # Skip if not a valid data row
            if not category_full or not subcategory:
                continue
                
            # Skip total rows (合計 Total)
            if "合計" in subcategory:
                continue
                
            # Extract category
            category = None
            for jp_cat, en_cat in self.brokerage_categories.items():
                if jp_cat in category_full:
                    category = en_cat
                    break
                    
            if not category:
                continue
                
            # Extract subcategory
            subcat = None
            for jp_sub, en_sub in self.subcategories.items():
                if jp_sub in subcategory:
                    subcat = en_sub
                    break
                    
            if not subcat:
                continue
                
            # Process VALUE
            if value is not None:
                value_str = self.handle_negative_values(str(value))
                if value_str:
                    value_code = f"JGBF.BROKERAGE_BREAKDOWN.{instrument_code}.TRADINGVALUE.{category}.{subcat}.VALUE.W"
                    value_desc = f"Trading by Type of Investors, {self.get_instrument_name(instrument_code)}, Breakdown of Brokerage, Trading Value, {category.title()}, {subcat.title()}, Value"
                    
                    results.append({
                        'code': value_code,
                        'description': value_desc,
                        'date': date_code,
                        'value': value_str
                    })
            
            # Process BALANCE
            if balance is not None:
                balance_str = self.handle_negative_values(str(balance))
                if balance_str:
                    balance_code = f"JGBF.BROKERAGE_BREAKDOWN.{instrument_code}.TRADINGVALUE.{category}.{subcat}.BALANCE.W"
                    balance_desc = f"Trading by Type of Investors, {self.get_instrument_name(instrument_code)}, Breakdown of Brokerage, Trading Value, {category.title()}, {subcat.title()}, Balance"
                    
                    results.append({
                        'code': balance_code,
                        'description': balance_desc,
                        'date': date_code, 
                        'value': balance_str
                    })
                    
        return results

    def get_instrument_name(self, instrument_code: str) -> str:
        """Get human-readable instrument name from code."""
        reverse_mapping = {v: k for k, v in self.instrument_mapping.items()}
        return reverse_mapping.get(instrument_code, instrument_code)

    def process_single_file(self, file_path: Path) -> List[Dict]:
        """Process a single Excel file."""
        logger.info(f"📊 Processing {file_path.name}")
        
        all_results = []
        date_code = self.extract_date_from_filename(file_path.stem)
        
        try:
            wb = load_workbook(file_path, data_only=True)
            
            # Process each relevant sheet
            for sheet_name in wb.sheetnames:
                if sheet_name == "Summary":
                    continue
                    
                # Determine if this is a table we want to process
                if "Table1_Main_Summary" in sheet_name:
                    table_type = "main_summary"
                elif "Table2_Brokerage_Bre" in sheet_name:
                    table_type = "brokerage_breakdown" 
                else:
                    continue  # Skip other tables for now
                    
                sheet_data = self.read_excel_sheet(file_path, sheet_name)
                if not sheet_data:
                    continue
                    
                # Extract instrument from subtitle
                instrument_code = self.extract_instrument_from_subtitle(sheet_data['subtitle'])
                if not instrument_code:
                    continue
                    
                # Parse based on table type
                if table_type == "main_summary":
                    results = self.parse_main_summary_table(sheet_data, instrument_code, date_code)
                elif table_type == "brokerage_breakdown":
                    results = self.parse_brokerage_breakdown_table(sheet_data, instrument_code, date_code)
                else:
                    continue
                    
                all_results.extend(results)
                logger.info(f"  • {sheet_name}: {len(results)} data points extracted")
                
        except Exception as e:
            logger.error(f"Error processing file {file_path.name}: {e}")
            
        return all_results

    def generate_output_file(self, all_data: List[Dict], output_filename: str):
        """Generate the final JGBF_DATA Excel file using the exact template structure."""
        if not all_data:
            logger.warning("No data to generate output file")
            return
            
        # Group data by time series code
        time_series_data = {}
        for item in all_data:
            code = item['code']
            if code not in time_series_data:
                time_series_data[code] = {
                    'description': item['description'],
                    'data_points': {}
                }
            time_series_data[code]['data_points'][item['date']] = item['value']
            
        # Get all unique dates and sort them
        all_dates = set()
        for ts_data in time_series_data.values():
            all_dates.update(ts_data['data_points'].keys())
        sorted_dates = sorted(list(all_dates))
        
        # Get the exact template structure
        template_columns = self.get_template_columns()
        
        # Create output workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "JGBF_DATA"
        
        # Set up headers using the exact template structure
        # Row 1: Time series codes
        ws['A1'] = "Date"
        for col_idx, column_def in enumerate(template_columns, start=2):
            ws.cell(row=1, column=col_idx, value=column_def['code'])
            
        # Row 2: Descriptions  
        ws['A2'] = "Description"
        for col_idx, column_def in enumerate(template_columns, start=2):
            ws.cell(row=2, column=col_idx, value=column_def['description'])
            
        # Data rows
        for row_idx, date in enumerate(sorted_dates, start=3):
            ws.cell(row=row_idx, column=1, value=date)
            
            for col_idx, column_def in enumerate(template_columns, start=2):
                code = column_def['code']
                if code in time_series_data:
                    value = time_series_data[code]['data_points'].get(date, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)
                else:
                    # Column exists in template but no data found
                    ws.cell(row=row_idx, column=col_idx, value="")
                
        # Apply formatting
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
        for cell in ws[2]:
            cell.font = header_font
            
        # Save file
        output_path = self.output_folder / output_filename
        wb.save(output_path)
        logger.info(f"💾 Output saved: {output_path}")
        logger.info(f"  • {len(template_columns)} template columns")
        logger.info(f"  • {len(sorted_dates)} date periods")
        logger.info(f"  • {len([col for col in template_columns if col['code'] in time_series_data])} columns with data")
        
        # Log which columns have data and which are empty
        columns_with_data = []
        columns_without_data = []
        for column_def in template_columns:
            if column_def['code'] in time_series_data:
                columns_with_data.append(column_def['code'])
            else:
                columns_without_data.append(column_def['code'])
                
        logger.info(f"📊 Data coverage:")
        logger.info(f"  • Columns with data: {len(columns_with_data)}")
        logger.info(f"  • Columns without data: {len(columns_without_data)}")
        
        if columns_without_data:
            logger.info("📝 Columns without data (will be empty):")
            for i, code in enumerate(columns_without_data[:5]):  # Show first 5
                logger.info(f"  - {code}")
            if len(columns_without_data) > 5:
                logger.info(f"  ... and {len(columns_without_data) - 5} more")

    def process_all_files(self):
        """Process all Excel files in the input folder."""
        if not self.input_folder.exists():
            logger.error(f"Input folder '{self.input_folder}' does not exist!")
            return
            
        # Skip Excel temporary files
        excel_files = [f for f in self.input_folder.glob("*_extracted.xlsx") if not f.name.startswith("~$")]
        if not excel_files:
            logger.error(f"No extracted Excel files found in '{self.input_folder}'")
            return
            
        logger.info(f"📁 Found {len(excel_files)} Excel files to process")
        
        all_data = []
        for file_path in excel_files:
            file_results = self.process_single_file(file_path)
            all_data.extend(file_results)
            
        if all_data:
            # Generate timestamp for output filename
            timestamp = datetime.now().strftime("%Y%m%d")
            output_filename = f"JGBF_DATA_{timestamp}.xlsx"
            
            self.generate_output_file(all_data, output_filename)
            
            print(f"\n🎉 PARSING COMPLETED!")
            print(f"  • Processed {len(excel_files)} input files")
            print(f"  • Extracted {len(all_data)} total data points")
            print(f"  • Output: {self.output_folder / output_filename}")
        else:
            logger.error("No data extracted from any files!")


def main():
    print("📊 JGBF Excel Parser - Fixed Version")
    print("=" * 55)
    print("🚀 Converts extracted Excel files to JGBF_DATA format")
    print("🔧 Includes fixes for subtitle mapping and complete template")
    print("=" * 55)
    
    parser = JGBFParser(
        input_folder="extracted_data",
        output_folder="parsed_output"
    )
    
    parser.process_all_files()


if __name__ == "__main__":
    main()